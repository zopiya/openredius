"""Report export builders: xlsx / pdf / csv (docs/03「报表统计」导出).

The full report bundles three modules — failure-reason distribution (summary),
endpoint-type mix, and per-department admission stats — into a multi-sheet
xlsx, a multi-section pdf, or (for csv, which is inherently single-table) the
department table kept from the original implementation.
"""

from __future__ import annotations

import csv
import io
from typing import Any


def _rows_from_summary(summary: dict[str, Any]) -> list[list[str]]:
    return [["label", "value"]] + [[f["label"], f["value"]] for f in summary.get("fail", [])]


def _rows_from_types(items: list[dict[str, Any]]) -> list[list[str]]:
    return [["label", "value"]] + [[r["label"], r["value"]] for r in items]


def _rows_from_departments(items: list[dict[str, Any]]) -> list[list[str]]:
    return [["部门", "在线", "成功", "失败", "成功率"]] + [
        [r["dept"], r["online"], r["ok"], r["fail"], r["rate"]] for r in items
    ]


def build_xlsx(
    summary: dict[str, Any],
    endpoint_types: list[dict[str, Any]],
    departments: list[dict[str, Any]],
    *,
    period: str,
) -> bytes:
    """Multi-sheet workbook; returns the raw xlsx bytes."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "失败原因分布"
    for row in _rows_from_summary(summary):
        ws.append(row)
    for row in _rows_from_types(endpoint_types):
        wb.create_sheet("终端类型").append(row)
    for row in _rows_from_departments(departments):
        wb.create_sheet("部门准入").append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf(
    summary: dict[str, Any],
    endpoint_types: list[dict[str, Any]],
    departments: list[dict[str, Any]],
    *,
    period: str,
) -> bytes:
    """Multi-section PDF using reportlab's built-in STSong-Light CJK font."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    base = getSampleStyleSheet()
    title_style = ParagraphStyle("cn-title", parent=base["Title"], fontName="STSong-Light")
    heading_style = ParagraphStyle("cn-heading", parent=base["Heading2"], fontName="STSong-Light")
    cell_style = ParagraphStyle("cn-cell", fontName="STSong-Light", fontSize=9, leading=12)

    def _table(title: str, rows: list[list[str]]) -> list[Any]:
        header = [Paragraph(str(c), cell_style) for c in rows[0]]
        body = [[Paragraph(str(c), cell_style) for c in r] for r in rows[1:]]
        table = Table([header, *body], hAlign="LEFT")
        table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f0f0f0")),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        return [Paragraph(title, heading_style), Spacer(1, 6), table, Spacer(1, 14)]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title=f"OpenRedius 准入报表 ({period})")
    story: list[Any] = [
        Paragraph("OpenRedius 准入报表", title_style),
        Paragraph(summary.get("sub", f"统计周期:{period}"), cell_style),
        Spacer(1, 12),
        *_table("失败原因分布", _rows_from_summary(summary)),
        *_table("终端类型占比", _rows_from_types(endpoint_types)),
        *_table("部门准入", _rows_from_departments(departments)),
    ]
    doc.build(story)
    return buf.getvalue()


def build_csv(departments: list[dict[str, Any]]) -> str:
    """Single-table department CSV (kept from the original export)."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    for row in _rows_from_departments(departments):
        writer.writerow(row)
    return buf.getvalue()
